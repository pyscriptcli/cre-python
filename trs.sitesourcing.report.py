import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter, range_boundaries
import re
import io
import zipfile
import requests
from copy import copy
import os

# --- PAGE CONFIGURATION ---
st.set_page_config(
    page_title="trs.sitesourcing.report",
    page_icon="",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# --- PROGRAMMATIC LIGHT MODE LOCK ---
_config_dir = ".streamlit"
_config_file = os.path.join(_config_dir, "config.toml")
if not os.path.exists(_config_file):
    os.makedirs(_config_dir, exist_ok=True)
    with open(_config_file, "w", encoding="utf-8") as f:
        f.write("[theme]\nbase=\"light\"\n")

# --- CUSTOM CSS ---
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Roboto:wght@300;400;500;700&display=swap');
    
    * {
        font-family: 'Roboto', 'Segoe UI', sans-serif !important;
    }
    
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}
    
    .main-header {
        display: none !important;
    }
    
    .block-container {
        padding-top: 0.3rem !important;
        padding-bottom: 3.5rem !important;
        max-width: 1200px !important;
    }
    
    .stButton > button {
        background-color: #003366 !important;
        color: white !important;
        border: none !important;
        border-radius: 3px !important;
        padding: 0.4rem 0.6rem !important;
        font-weight: 500 !important;
        font-size: 0.8rem !important;
        transition: all 0.2s ease;
        width: 100%;
        min-height: 36px;
        opacity: 1 !important;
        visibility: visible !important;
    }
    .stButton > button:hover {
        background-color: #002244 !important;
        box-shadow: 0 2px 6px rgba(0, 51, 102, 0.3);
    }
    .stButton > button:active {
        background-color: #001a33 !important;
    }
    
    .stDownloadButton > button {
        background-color: #28a745 !important;
        color: white !important;
        border: none !important;
        border-radius: 3px !important;
        padding: 0.4rem 0.6rem !important;
        font-weight: 500 !important;
        font-size: 0.8rem !important;
        width: 100%;
        min-height: 36px;
        opacity: 1 !important;
        visibility: visible !important;
    }
    .stDownloadButton > button:hover {
        background-color: #218838 !important;
        box-shadow: 0 2px 6px rgba(40, 167, 69, 0.3);
    }
    
    .stButton button p, .stButton button span {
        color: white !important;
        opacity: 1 !important;
    }
    .stDownloadButton button p, .stDownloadButton button span {
        color: white !important;
        opacity: 1 !important;
    }
    
    div[data-testid="stContainer"] {
        background-color: white !important;
        border-radius: 4px;
        padding: 0.6rem;
        box-shadow: 0 1px 3px rgba(0,0,0,0.1);
        border: 1px solid #e0e0e0;
    }
    
    h1, h2, h3, h4, h5, h6, p, label, span, div {
        color: #003366 !important;
    }
    
    .stCheckbox label {
        font-weight: 400;
        color: #003366 !important;
        font-size: 0.8rem;
    }
    .stCheckbox label:hover {
        color: #002244 !important;
    }
    .stCheckbox {
        margin-bottom: 0.1rem;
    }
    .stCheckbox > div {
        background-color: white !important;
    }
    
    .stProgress > div > div {
        background-color: #003366 !important;
    }
    
    hr {
        border-color: #003366 !important;
        opacity: 0.15;
        margin: 0.4rem 0;
    }
    
    .metric-card {
        background-color: white !important;
        border-radius: 4px;
        padding: 0.4rem;
        box-shadow: 0 1px 3px rgba(0,0,0,0.1);
        border-left: 3px solid #003366;
        text-align: center;
        margin-bottom: 0.3rem;
    }
    .metric-value {
        font-size: 1.1rem;
        font-weight: 500;
        color: #003366 !important;
        font-family: 'Roboto', sans-serif;
    }
    .metric-label {
        color: #003366 !important;
        font-size: 0.6rem;
        font-weight: 400;
        text-transform: uppercase;
        letter-spacing: 0.5px;
        font-family: 'Roboto', sans-serif;
    }
    
    .checkbox-container {
        max-height: 280px;
        overflow-y: auto;
        padding-right: 4px;
        background-color: white !important;
    }
    .checkbox-container::-webkit-scrollbar {
        width: 4px;
    }
    .checkbox-container::-webkit-scrollbar-track {
        background: #f1f1f1;
        border-radius: 2px;
    }
    .checkbox-container::-webkit-scrollbar-thumb {
        background: #003366;
        border-radius: 2px;
    }
    
    .stAlert {
        border-radius: 4px;
        padding: 0.4rem;
    }
    .stAlert[data-baseweb="notification"] {
        border-left-color: #003366 !important;
    }
    
    .stSuccess {
        background-color: #d4edda !important;
        color: #003366 !important;
    }
    .stSuccess * {
        color: #003366 !important;
    }
    
    .stWarning {
        background-color: #fff3cd !important;
        color: #003366 !important;
    }
    .stWarning * {
        color: #003366 !important;
    }
    
    .stError {
        background-color: #f8d7da !important;
        color: #003366 !important;
    }
    .stError * {
        color: #003366 !important;
    }
    
    .stInfo {
        background-color: #e8f0fe !important;
        color: #003366 !important;
    }
    .stInfo * {
        color: #003366 !important;
    }
    
    .stSpinner > div {
        border-color: #003366 !important;
    }
    
    .stMarkdown, .stMarkdown * {
        color: #003366 !important;
    }
    
    .footer-overlay {
        position: fixed !important;
        bottom: 0 !important;
        left: 0 !important;
        right: 0 !important;
        height: 15px !important;
        background-color: white !important;
        z-index: 9999999 !important;
        box-shadow: 0 -2px 10px rgba(0,0,0,0.05) !important;
        border-top: 1px solid #e8e8e8 !important;
        pointer-events: none !important;
    }
    
    .main {
        padding-bottom: 30px !important;
    }
    
    .stApp {
        padding-bottom: 15px !important;
    }
    
    .stAppViewContainer {
        padding-bottom: 15px !important;
    }
</style>
""", unsafe_allow_html=True)

# --- PERSISTENT FOOTER ---
st.markdown("""
<div class="footer-overlay"></div>
""", unsafe_allow_html=True)

# --- CONFIGURATION ---
SOURCE_URL = "https://docs.google.com/spreadsheets/d/14nhO9u7zJRcOoux8I7l2IzwU7iQZNW9fRX6TCip47CE/export?format=xlsx"
TEMPLATE_URL = "https://docs.google.com/spreadsheets/d/1uS3xmnPi0o4c_EayQtURYDSMMPRDRGSb/export?format=xlsx"

# --- HELPER FUNCTIONS ---
def download_file(url):
    try:
        response = requests.get(url, timeout=30)
        if response.status_code == 200:
            return io.BytesIO(response.content)
        return None
    except:
        return None

def get_placeholders(sheet):
    placeholders = set()
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                matches = re.findall(r"\{\{(.*?)\}\}", cell.value)
                for match in matches:
                    name = match.split(":")[0].strip() if ":" in match else match.strip()
                    placeholders.add(name)
    return sorted(list(placeholders))

def sanitize_tab_name(name, existing_names):
    illegal_chars = r'[\\/*?\[\]:]'
    clean_name = re.sub(illegal_chars, '', str(name))
    base_name = clean_name[:31]
    if base_name not in existing_names:
        existing_names.add(base_name)
        return base_name
    counter = 1
    while True:
        suffix = f" ({counter})"
        max_len = 31 - len(suffix)
        new_name = f"{clean_name[:max_len]}{suffix}"
        if new_name not in existing_names:
            existing_names.add(new_name)
            return new_name
        counter += 1

def clone_cell_styles(source_cell, target_cell):
    if source_cell.font:
        target_cell.font = Font(
            name=source_cell.font.name,
            size=source_cell.font.size,
            bold=source_cell.font.bold,
            italic=source_cell.font.italic,
            color=copy(source_cell.font.color),
            underline=source_cell.font.underline,
            strike=source_cell.font.strike,
            vertAlign=source_cell.font.vertAlign,
            scheme=source_cell.font.scheme
        )
    if source_cell.alignment:
        target_cell.alignment = Alignment(
            horizontal=source_cell.alignment.horizontal,
            vertical=source_cell.alignment.vertical,
            text_rotation=source_cell.alignment.text_rotation,
            wrap_text=source_cell.alignment.wrap_text,
            shrink_to_fit=source_cell.alignment.shrink_to_fit,
            indent=source_cell.alignment.indent
        )
    if source_cell.border:
        target_cell.border = Border(
            left=copy(source_cell.border.left),
            right=copy(source_cell.border.right),
            top=copy(source_cell.border.top),
            bottom=copy(source_cell.border.bottom),
            diagonal=copy(source_cell.border.diagonal),
            diagonal_direction=source_cell.border.diagonal_direction,
            outline=copy(source_cell.border.outline),
            vertical=copy(source_cell.border.vertical),
            horizontal=copy(source_cell.border.horizontal)
        )
    if source_cell.fill:
        target_cell.fill = copy(source_cell.fill)

def copy_and_merge_aware_injection(template_ws, target_ws, coord, data_value):
    if not target_ws:
        return
    target_cell = target_ws[coord]
    template_cell = template_ws[coord]
    target_cell.value = data_value
    clone_cell_styles(template_cell, target_cell)
    merged_range_string = None
    for merged_range in template_ws.merged_cells.ranges:
        if template_cell.coordinate in merged_range:
            merged_range_string = str(merged_range)
            break
    if merged_range_string:
        min_col, min_row, max_col, max_row = range_boundaries(merged_range_string)
        overlapping_target_ranges = []
        for target_range in target_ws.merged_cells.ranges:
            t_min_c, t_min_r, t_max_c, t_max_r = range_boundaries(str(target_range))
            if not (max_col < t_min_c or min_col > t_max_c or max_row < t_min_r or min_row > t_max_r):
                overlapping_target_ranges.append(target_range)
        for target_conflict in overlapping_target_ranges:
            try:
                target_ws.unmerge_cells(str(target_conflict))
            except:
                pass
        try:
            target_ws.merge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)
        except:
            pass
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                sub_coord = f"{get_column_letter(c)}{r}"
                if sub_coord != coord:
                    clone_cell_styles(template_ws[sub_coord], target_ws[sub_coord])

# --- LOAD FILES ---
def load_files():
    source_data = download_file(SOURCE_URL)
    template_data = download_file(TEMPLATE_URL)
    return source_data, template_data

# --- LOAD DATA ---
with st.spinner("Loading..."):
    source_data, template_data = load_files()

if source_data is None or template_data is None:
    st.error("Failed to load files. Make sure they are publicly accessible.")
    st.stop()

# --- PROCESS DATA ---
df = pd.read_excel(source_data)
df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
df.columns = df.columns.str.strip().str.upper()

if "TRADE AREA" not in df.columns or "SITE NAME" not in df.columns:
    st.error("Data must contain 'TRADE AREA' and 'SITE NAME' columns.")
    st.stop()

# Check for TRADE AREA NO column for unique count
trade_area_no_col = None
for col in df.columns:
    if "TRADE AREA NO" in col.upper() or "TRADE AREA #" in col.upper():
        trade_area_no_col = col
        break

# Get unique trade area count from TRADE AREA NO (for stats)
if trade_area_no_col:
    unique_trade_areas_count = len(df[trade_area_no_col].dropna().unique())
else:
    unique_trade_areas_count = len(df["TRADE AREA"].unique())

# Get ALL unique TRADE AREA names for checkboxes (not connected to TRADE AREA NO)
all_trade_areas = sorted(df["TRADE AREA"].dropna().unique())

# Create a mapping of TRADE AREA NO to TRADE AREA name for filtering when generating
trade_area_no_to_name = {}
if trade_area_no_col:
    unique_combos = df[[trade_area_no_col, "TRADE AREA"]].drop_duplicates()
    for _, row in unique_combos.iterrows():
        key = str(row[trade_area_no_col])
        value = str(row["TRADE AREA"])
        trade_area_no_to_name[key] = value

template_wb = openpyxl.load_workbook(template_data)
template_sheet = template_wb.active
placeholders = get_placeholders(template_sheet)

if not placeholders:
    st.warning("No placeholders found in template.")
    st.stop()

# --- SESSION STATE ---
if 'zip_data' not in st.session_state:
    st.session_state.zip_data = None
if 'selected_tas' not in st.session_state:
    st.session_state.selected_tas = []
if 'single_file' not in st.session_state:
    st.session_state.single_file = None

# --- 3-COLUMN LAYOUT ---
col1, col2, col3 = st.columns([0.8, 1.4, 0.8])

# --- COLUMN 1: METRICS ---
with col1:
    st.markdown("### Stats")
    
    st.markdown(f"""
    <div class="metric-card">
        <div class="metric-value">{len(df)}</div>
        <div class="metric-label">Records</div>
    </div>
    <div class="metric-card">
        <div class="metric-value">{unique_trade_areas_count}</div>
        <div class="metric-label">Trade Areas</div>
    </div>
    <div class="metric-card">
        <div class="metric-value">{len(placeholders)}</div>
        <div class="metric-label">Placeholders</div>
    </div>
    """, unsafe_allow_html=True)
    
    # Refresh button under Placeholders
    st.markdown("---")
    if st.button("Refresh Data", use_container_width=True):
        st.cache_data.clear()
        st.cache_resource.clear()
        if os.path.exists(_config_file):
            os.remove(_config_file)
        st.rerun()

# --- COLUMN 2: TRADE AREA SELECTION ---
with col2:
    st.markdown("### Select Trade Areas")
    
    # Select All / Clear All buttons
    btn_col1, btn_col2 = st.columns(2)
    with btn_col1:
        if st.button("Select All", use_container_width=True):
            for ta in all_trade_areas:
                st.session_state[f"ta_{ta}"] = True
            st.rerun()
    with btn_col2:
        if st.button("Clear All", use_container_width=True):
            for ta in all_trade_areas:
                st.session_state[f"ta_{ta}"] = False
            st.rerun()
    
    # Initialize session state for checkboxes - default unchecked
    for ta in all_trade_areas:
        if f"ta_{ta}" not in st.session_state:
            st.session_state[f"ta_{ta}"] = False
    
    # Checkboxes in scrollable container - display ALL TRADE AREA names
    st.markdown('<div class="checkbox-container">', unsafe_allow_html=True)
    for ta in all_trade_areas:
        st.checkbox(ta, key=f"ta_{ta}")
    st.markdown('</div>', unsafe_allow_html=True)

# --- COLUMN 3: ACTIONS ---
with col3:
    st.markdown("### Actions")
    
    if st.session_state.zip_data is None and st.session_state.single_file is None:
        if st.button("Generate Reports", use_container_width=True, type="primary"):
            # Get selected trade area names from session state
            selected_trade_areas = [ta for ta in all_trade_areas if st.session_state.get(f"ta_{ta}", False)]
            
            if not selected_trade_areas:
                st.warning("Select at least one Trade Area.")
            else:
                with st.spinner("Generating..."):
                    progress_bar = st.progress(0)
                    
                    # Filter based on selected trade area names
                    filtered_df = df[df["TRADE AREA"].isin(selected_trade_areas)]
                    
                    groups = filtered_df.groupby("TRADE AREA")
                    total_groups = len(groups)
                    
                    # If only one trade area selected, create single file
                    if len(selected_trade_areas) == 1:
                        selected_ta = selected_trade_areas[0]
                        group = filtered_df[filtered_df["TRADE AREA"] == selected_ta]
                        
                        template_data.seek(0)
                        wb = openpyxl.load_workbook(template_data)
                        base_sheet = wb.active
                        base_sheet.title = "TEMPLATE_TO_DELETE"
                        existing_tabs = set()
                        
                        for _, row in group.iterrows():
                            site_name = row.get("SITE NAME", "Unknown")
                            safe_tab_name = sanitize_tab_name(site_name, existing_tabs)
                            new_sheet = wb.copy_worksheet(base_sheet)
                            new_sheet.title = safe_tab_name
                            
                            for row_cells in new_sheet.iter_rows():
                                for cell in row_cells:
                                    if isinstance(cell.value, str) and "{{" in cell.value:
                                        new_val = cell.value
                                        has_injected = False
                                        for ph in placeholders:
                                            target_regex = r"\{\{\s*" + re.escape(ph) + r"(\s*:.*?)?\}\}"
                                            if re.search(target_regex, new_val):
                                                raw_data_val = row.get(ph.upper(), "")
                                                if pd.isna(raw_data_val) or raw_data_val is None:
                                                    raw_data_val = ""
                                                val_str = str(raw_data_val)
                                                new_val = re.sub(target_regex, val_str, new_val)
                                                if val_str.strip() != "":
                                                    has_injected = True
                                        if has_injected and new_val != "":
                                            copy_and_merge_aware_injection(base_sheet, new_sheet, cell.coordinate, new_val.strip())
                                        elif new_val == "":
                                            cell.value = ""
                                        else:
                                            cell.value = new_val.strip() if new_val else ""
                        
                        wb.remove(base_sheet)
                        wb_buffer = io.BytesIO()
                        wb.save(wb_buffer)
                        safe_filename = str(selected_ta).replace("/", "-").replace("\\", "-")
                        
                        st.session_state.single_file = {
                            "data": wb_buffer.getvalue(),
                            "name": f"{safe_filename}.xlsx"
                        }
                        progress_bar.progress(1.0)
                    
                    else:
                        # Multiple trade areas - create zip
                        zip_buffer = io.BytesIO()
                        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
                            for i, (trade_area, group) in enumerate(groups):
                                template_data.seek(0)
                                wb = openpyxl.load_workbook(template_data)
                                base_sheet = wb.active
                                base_sheet.title = "TEMPLATE_TO_DELETE"
                                existing_tabs = set()
                                
                                for _, row in group.iterrows():
                                    site_name = row.get("SITE NAME", "Unknown")
                                    safe_tab_name = sanitize_tab_name(site_name, existing_tabs)
                                    new_sheet = wb.copy_worksheet(base_sheet)
                                    new_sheet.title = safe_tab_name
                                    
                                    for row_cells in new_sheet.iter_rows():
                                        for cell in row_cells:
                                            if isinstance(cell.value, str) and "{{" in cell.value:
                                                new_val = cell.value
                                                has_injected = False
                                                for ph in placeholders:
                                                    target_regex = r"\{\{\s*" + re.escape(ph) + r"(\s*:.*?)?\}\}"
                                                    if re.search(target_regex, new_val):
                                                        raw_data_val = row.get(ph.upper(), "")
                                                        if pd.isna(raw_data_val) or raw_data_val is None:
                                                            raw_data_val = ""
                                                        val_str = str(raw_data_val)
                                                        new_val = re.sub(target_regex, val_str, new_val)
                                                        if val_str.strip() != "":
                                                            has_injected = True
                                                if has_injected and new_val != "":
                                                    copy_and_merge_aware_injection(base_sheet, new_sheet, cell.coordinate, new_val.strip())
                                                elif new_val == "":
                                                    cell.value = ""
                                                else:
                                                    cell.value = new_val.strip() if new_val else ""
                                
                                wb.remove(base_sheet)
                                wb_buffer = io.BytesIO()
                                wb.save(wb_buffer)
                                safe_filename = str(trade_area).replace("/", "-").replace("\\", "-")
                                zip_file.writestr(f"{safe_filename}.xlsx", wb_buffer.getvalue())
                                progress_bar.progress((i + 1) / total_groups)
                        
                        st.session_state.zip_data = zip_buffer.getvalue()
                    
                    st.rerun()
    
    # Show download buttons
    if st.session_state.single_file is not None:
        st.success("Ready")
        st.download_button(
            "Download (.xlsx)",
            data=st.session_state.single_file["data"],
            file_name=st.session_state.single_file["name"],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
        if st.button("Reset", use_container_width=True):
            st.session_state.single_file = None
            st.rerun()
    
    if st.session_state.zip_data is not None:
        st.success("Ready")
        st.download_button(
            "Download (.zip)",
            data=st.session_state.zip_data,
            file_name="Reports.zip",
            mime="application/zip",
            use_container_width=True
        )
        if st.button("Reset", use_container_width=True):
            st.session_state.zip_data = None
            st.rerun()
