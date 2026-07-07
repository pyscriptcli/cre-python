import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter, range_boundaries
import re
import io
import requests
from copy import copy
import os
import hashlib
from openpyxl import load_workbook
from datetime import datetime

# --- PAGE CONFIGURATION ---
st.set_page_config(
    page_title="trs.sitesourcing.viewer",
    page_icon="",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# --- PROGRAMMATIC LIGHT MODE LOCK ---
_config_dir = ".streamlit"
_config_file = os.path.join(_config_dir, "config.toml")
if not os.path.exists(_config_file):
    os.makedirs(_config_dir, exist_ok=True)
    with open(_config_file, "w", encoding="utf-8") as f:
        f.write("[theme]\nbase=\"light\"\n")

# --- CUSTOM CSS ---
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Roboto:wght@300;400;500;700&display=swap');
    * { font-family: 'Roboto', 'Segoe UI', sans-serif !important; }
    #MainMenu {visibility: hidden !important;}
    footer {visibility: hidden !important;}
    header {visibility: hidden !important;}
    button[title="View source"] {display: none !important;}
    .stAppDeployButton {display: none !important;}
    div[data-testid="stStatusWidget"] {display: none !important;}
    
    .block-container {
        padding-top: 0.5rem !important;
        padding-bottom: 0.5rem !important;
        padding-left: 0.5rem !important;
        padding-right: 0.5rem !important;
        max-width: 100% !important;
    }
    .stButton > button, .stDownloadButton > button {
        background-color: #e8e8e8 !important;
        color: #333333 !important;
        border: 1px solid #d0d0d0 !important;
        border-radius: 2px !important;
        padding: 0.1rem 0.3rem !important;
        font-size: 0.7rem !important;
        min-height: 24px !important;
        height: 24px !important;
        width: 100%;
    }
    .stSelectbox label { 
        font-size: 0.75rem !important;
        font-weight: 500 !important;
        color: #333333 !important;
        margin-bottom: 2px !important;
    }
    .stSelectbox > div > div {
        background-color: #fafafa !important;
        border-color: #d0d0d0 !important;
        min-height: 24px !important;
        height: 24px !important;
    }
    .stSelectbox > div > div > div { padding-top: 0 !important; padding-bottom: 0 !important; font-size: 0.7rem !important; }
    div[data-testid="stHorizontalBlock"] { gap: 0.5rem !important; align-items: flex-end !important; }
    
    /* Document Display Container */
    .excel-container {
        background-color: #ffffff !important;
        border-radius: 2px;
        padding: 0.4rem;
        border: 1px solid #d0d0d0;
        overflow: auto;
        margin-top: 0.2rem;
        width: 100%;
    }
    .excel-container table {
        border-collapse: collapse;
        width: 100%;
        font-size: 10px;
        table-layout: fixed;
    }
    .excel-container td {
        padding: 4px 6px;
        word-break: break-word !important;
        white-space: normal !important;
        vertical-align: middle;
    }
    
    /* Tab headers styling */
    button[data-baseweb="tab"] {
        font-size: 0.85rem !important;
        font-weight: 600 !important;
        padding: 6px 16px !important;
        color: #555555 !important;
    }
    button[data-baseweb="tab"][aria-selected="true"] {
        color: #800000 !important;
        border-bottom-color: #800000 !important;
    }
    
    /* Card layout panels for assets */
    .asset-card {
        border: 1px solid #e0e0e0;
        border-radius: 4px;
        padding: 10px;
        background-color: #fafafa;
        margin-bottom: 15px;
        text-align: center;
        min-height: 320px;
    }
    .asset-title {
        font-size: 0.75rem;
        font-weight: bold;
        color: #333333;
        margin-top: 6px;
    }
    .photo-grid {
        display: grid;
        grid-template-columns: repeat(auto-fit, minmax(300px, 1fr));
        gap: 20px;
    }
    .photo-item {
        border: 1px solid #e0e0e0;
        border-radius: 4px;
        padding: 10px;
        background-color: #fafafa;
        text-align: center;
    }
    .photo-item img {
        max-width: 100%;
        max-height: 350px;
        object-fit: contain;
        border-radius: 2px;
    }
    
    /* Password visibility toggle */
    .password-toggle {
        position: relative;
    }
    .password-toggle .eye-icon {
        position: absolute;
        right: 10px;
        top: 50%;
        transform: translateY(-50%);
        cursor: pointer;
        color: #666;
        font-size: 1.2rem;
    }
</style>
""", unsafe_allow_html=True)

# --- LOGIN VERIFICATION LOGIC ---
TARGET_HASH = "6e7dfba0b39da481db37c3263c61cac6"
if 'authenticated' not in st.session_state:
    st.session_state.authenticated = False
if 'show_password' not in st.session_state:
    st.session_state.show_password = False

def check_password(password):
    return hashlib.md5(password.encode('utf-8')).hexdigest() == TARGET_HASH

if not st.session_state.authenticated:
    r1_col1, r1_col2, r1_col3 = st.columns([1, 1.2, 1])
    with r1_col2:
        st.markdown("<h3 style='text-align: center; margin-top:50px;'>Access Required</h3>", unsafe_allow_html=True)
        
        col_pwd, col_btn = st.columns([4, 1])
        with col_pwd:
            if st.session_state.show_password:
                password_input = st.text_input("Enter password:", type="default", label_visibility="collapsed")
            else:
                password_input = st.text_input("Enter password:", type="password", label_visibility="collapsed")
        with col_btn:
            eye_icon = "👁️" if st.session_state.show_password else "👁️‍🗨️"
            if st.button(eye_icon, key="toggle_pwd", help="Toggle password visibility"):
                st.session_state.show_password = not st.session_state.show_password
                st.rerun()
        
        if st.button("Login", use_container_width=True) or password_input:
            if check_password(password_input):
                st.session_state.authenticated = True
                st.rerun()
            else:
                st.error("Invalid token string provided.")
    st.stop()

# --- CONFIGURATION ---
SOURCE_URL = "https://docs.google.com/spreadsheets/d/14nhO9u7zJRcOoux8I7l2IzwU7iQZNW9fRX6TCip47CE/export?format=xlsx"
TEMPLATE_URL = "https://docs.google.com/spreadsheets/d/1uS3xmnPi0o4c_EayQtURYDSMMPRDRGSb/export?format=xlsx"
DRIVE_FOLDER_URL = "https://drive.google.com/drive/folders/13sLmXzxQvV12_ypTBRG2QW1yVIHaanba"

# --- HELPER FUNCTIONS ---
@st.cache_data(ttl=3600)
def download_file(url):
    try:
        response = requests.get(url, timeout=30)
        return io.BytesIO(response.content) if response.status_code == 200 else None
    except:
        return None

def get_placeholders(sheet):
    placeholders = set()
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                matches = re.findall(r"\{\{(.*?)\}\}", cell.value)
                for match in matches:
                    name = match.split(":")[0].strip() if ":" in match else match.strip()
                    placeholders.add(name)
    return sorted(list(placeholders))

def clone_cell_styles(source_cell, target_cell):
    if source_cell.font:
        target_cell.font = Font(
            name=source_cell.font.name, size=source_cell.font.size, bold=source_cell.font.bold,
            italic=source_cell.font.italic, color=copy(source_cell.font.color)
        )
    if source_cell.alignment:
        target_cell.alignment = Alignment(horizontal=source_cell.alignment.horizontal, vertical=source_cell.alignment.vertical, wrap_text=True)
    if source_cell.fill:
        target_cell.fill = copy(source_cell.fill)
    if source_cell.border:
        target_cell.border = copy(source_cell.border)

def copy_and_merge_aware_injection(template_ws, target_ws, coord, data_value):
    if not target_ws: return
    target_cell = target_ws[coord]
    template_cell = template_ws[coord]
    target_cell.value = data_value
    clone_cell_styles(template_cell, target_cell)

def sanitize_tab_name(name, existing_names):
    clean_name = re.sub(r'[\\/*?\[\]:]', '', str(name))[:31]
    if clean_name not in existing_names:
        existing_names.add(clean_name)
        return clean_name
    counter = 1
    while True:
        new_name = f"{clean_name[:27]} ({counter})"
        if new_name not in existing_names:
            existing_names.add(new_name)
            return new_name
        counter += 1

def transform_to_direct_download(drive_url):
    url_str = str(drive_url).strip()
    if "drive.google.com" in url_str:
        file_id_match = re.search(r'/d/([a-zA-Z0-9-_]+)', url_str)
        if file_id_match:
            return f"https://drive.google.com/uc?export=download&id={file_id_match.group(1)}"
        id_param_match = re.search(r'id=([a-zA-Z0-9-_]+)', url_str)
        if id_param_match:
            return f"https://drive.google.com/uc?export=download&id={id_param_match.group(1)}"
    return url_str

def extract_file_id_from_drive_url(url):
    """Extract file ID from various Google Drive URL formats"""
    url_str = str(url).strip()
    
    # Check for /d/ pattern
    file_id_match = re.search(r'/d/([a-zA-Z0-9-_]+)', url_str)
    if file_id_match:
        return file_id_match.group(1)
    
    # Check for id= pattern
    id_param_match = re.search(r'id=([a-zA-Z0-9-_]+)', url_str)
    if id_param_match:
        return id_param_match.group(1)
    
    # Check for file/d/ pattern
    file_id_match = re.search(r'file/d/([a-zA-Z0-9-_]+)', url_str)
    if file_id_match:
        return file_id_match.group(1)
    
    return None

def parse_link_cell(cell_value):
    if pd.isna(cell_value) or not str(cell_value).strip():
        return []
    return [url.strip() for url in str(cell_value).split(",") if url.strip()]

def get_today_date():
    return datetime.now().strftime('%B %d, %Y')

def process_value(val, is_export=False):
    """Process a value for display, handling NaN, None, and formatting"""
    if pd.isna(val) or val is None:
        return ""
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    if hasattr(val, 'strftime'):
        return val.strftime('%B %d, %Y')
    return str(val).strip()

# --- HTML TEMPLATE BLUEPRINT DEFINITION ---
RAW_TEMPLATE_HTML = """
<style type="text/css">
    .ritz .waffle a { color: inherit; }
    .ritz .waffle .s25{border-right:none;border-bottom:1px SOLID transparent;background-color:#ffffff;text-align:left;color:#000000;font-family:Arial;font-size:12pt;vertical-align:middle;white-space:nowrap;direction:ltr;padding:0px 3px 0px 3px;}
    .ritz .waffle .s20{border-bottom:1px SOLID #000000;border-right:1px SOLID transparent;background-color:#b7b7b7;text-align:left;color:#000000;font-family:Arial;font-size:10pt;vertical-align:middle;white-space:nowrap;direction:ltr;padding:0px 3px 0px 3px;}
    .ritz .waffle .s15{border-bottom:1px SOLID transparent;border-right:1px SOLID transparent;background-color:#b7b7b7;text-align:left;color:#000000;font-family:Arial;font-size:10pt;vertical-align:middle;white-space:nowrap;direction:ltr;padding:0px 3px 0px 3px;}
    .ritz .waffle .s1{border-bottom:1px SOLID #bfbfbf;border-right:1px SOLID #bfbfbf;background-color:#ffffff;text-align:left;font-weight:bold;color:#000000;font-family:Arial;font-size:12pt;vertical-align:middle;white-space:nowrap;direction:ltr;padding:0px 3px 0px 3px;}
    .ritz .waffle .s10{background-color:#bfbfbf;text-align:left;color:#000000;font-family:Arial;font-size:12pt;vertical-align:middle;white-space:nowrap;direction:ltr;padding:0px 3px 0px 3px;}
    .ritz .waffle .s24{background-color:#ffffff;text-align:left;font-weight:bold;color:#000000;font-family:Arial;font-size:12pt;vertical-align:middle;white-space:nowrap;direction:ltr;padding:0px 3px 0px 3px;}
    .ritz .waffle .s6{border-bottom:1px SOLID #bfbfbf;background-color:#ffffff;text-align:left;color:#000000;font-family:Arial;font-size:12pt;vertical-align:middle;white-space:nowrap;direction:ltr;padding:0px 3px 0px 3px;}
    .ritz .waffle .s27{border-left:none;background-color:#ffffff;text-align:left;color:#000000;font-family:Arial;font-size:12pt;vertical-align:middle;white-space:nowrap;direction:ltr;padding:0px 3px 0px 3px;}
    .ritz .waffle .s22{border-bottom:1px SOLID #000000;background-color:#b7b7b7;text-align:left;color:#000000;font-family:Arial;font-size:10pt;vertical-align:middle;white-space:nowrap;direction:ltr;padding:0px 3px 0px 3px;}
    .ritz .waffle .s13{background-color:#b7b7b7;text-align:left;font-weight:bold;color:#ff0000;font-family:Arial;font-size:10pt;vertical-align:middle;white-space:nowrap;direction:ltr;padding:0px 3px 0px 3px;}
    .ritz .waffle .s17{background-color:#b7b7b7;text-align:left;color:#000000;font-family:Arial;font-size:10pt;vertical-align:middle;white-space:nowrap;direction:ltr;padding:0px 3px 0px 3px;}
    .ritz .waffle .s18{border-bottom:1px SOLID transparent;border-right:1px SOLID #bfbfbf;background-color:#b7b7b7;text-align:left;color:#ff0000;font-family:Arial;font-size:10pt;vertical-align:middle;white-space:nowrap;direction:ltr;padding:0px 3px 0px 3px;}
    .ritz .waffle .s4{border-bottom:1px SOLID transparent;border-right:1px SOLID transparent;background-color:#bfbfbf;text-align:left;color:#000000;font-family:Arial;font-size:12pt;vertical-align:middle;white-space:nowrap;direction:ltr;padding:0px 3px 0px 3px;}
    .ritz .waffle .s0{border-bottom:1px SOLID #bfbfbf;border-right:1px SOLID #bfbfbf;background-color:#800000;text-align:center;font-weight:bold;color:#ffffff;font-family:Arial;font-size:12pt;vertical-align:middle;white-space:nowrap;direction:ltr;padding:0px 3px 0px 3px;}
    .ritz .waffle .s8{border-bottom:1px SOLID transparent;border-right:1px SOLID transparent;background-color:#ffffff;text-align:left;color:#ff0000;font-family:Arial;font-size:12pt;vertical-align:middle;white-space:nowrap;direction:ltr;padding:0px 3px 0px 3px;}
    .ritz .waffle .s19{border-bottom:1px SOLID #000000;background-color:#b7b7b7;text-align:left;color:#ff0000;font-family:Arial;font-size:10pt;vertical-align:middle;white-space:nowrap;direction:ltr;padding:0px 3px 0px 3px;}
    .ritz .waffle .s2{background-color:#ffffff;text-align:left;color:#000000;font-family:Arial;font-size:12pt;vertical-align:middle;white-space:nowrap;direction:ltr;padding:0px 3px 0px 3px;}
    .ritz .waffle .s11{border-bottom:1px SOLID #000000;background-color:#ffffff;text-align:left;color:#000000;font-family:Arial;font-size:12pt;vertical-align:middle;white-space:nowrap;direction:ltr;padding:0px 3px 0px 3px;}
    .ritz .waffle .s21{border-bottom:1px SOLID #000000;border-right:1px SOLID transparent;background-color:#b7b7b7;text-align:left;color:#ff0000;font-family:Arial;font-size:10pt;vertical-align:middle;white-space:nowrap;direction:ltr;padding:0px 3px 0px 3px;}
    .ritz .waffle .s14{background-color:#b7b7b7;text-align:left;color:#ff0000;font-family:Arial;font-size:10pt;vertical-align:middle;white-space:nowrap;direction:ltr;padding:0px 3px 0px 3px;}
    .ritz .waffle .s3{border-right:1px SOLID #bfbfbf;background-color:#ffffff;text-align:left;color:#000000;font-family:Arial;font-size:12pt;vertical-align:middle;white-space:nowrap;direction:ltr;padding:0px 3px 0px 3px;}
    .ritz .waffle .s23{border-bottom:1px SOLID #000000;border-right:1px SOLID #bfbfbf;background-color:#b7b7b7;text-align:left;color:#ff0000;font-family:Arial;font-size:10pt;vertical-align:middle;white-space:nowrap;direction:ltr;padding:0px 3px 0px 3px;}
    .ritz .waffle .s5{border-bottom:1px SOLID transparent;border-right:1px SOLID transparent;background-color:#ffffff;text-align:left;color:#000000;font-family:Arial;font-size:12pt;vertical-align:middle;white-space:nowrap;direction:ltr;padding:0px 3px 0px 3px;}
    .ritz .waffle .s9{border-bottom:1px SOLID transparent;border-right:1px SOLID #bfbfbf;background-color:#bfbfbf;text-align:left;color:#000000;font-family:Arial;font-size:12pt;vertical-align:middle;white-space:nowrap;direction:ltr;padding:0px 3px 0px 3px;}
    .ritz .waffle .s12{border-bottom:1px SOLID #000000;border-right:1px SOLID #bfbfbf;background-color:#ffffff;text-align:left;color:#000000;font-family:Arial;font-size:12pt;vertical-align:middle;white-space:nowrap;direction:ltr;padding:0px 3px 0px 3px;}
    .ritz .waffle .s16{border-bottom:1px SOLID transparent;border-right:1px SOLID transparent;background-color:#b7b7b7;text-align:left;color:#ff0000;font-family:Arial;font-size:10pt;vertical-align:middle;white-space:nowrap;direction:ltr;padding:0px 3px 0px 3px;}
    .ritz .waffle .s26{border-left:none;border-bottom:1px SOLID transparent;background-color:#ffffff;text-align:left;color:#000000;font-family:Arial;font-size:12pt;vertical-align:middle;white-space:nowrap;direction:ltr;padding:0px 3px 0px 3px;}
    .ritz .waffle .s7{border-bottom:1px SOLID #bfbfbf;border-right:1px SOLID #bfbfbf;background-color:#ffffff;text-align:left;color:#000000;font-family:Arial;font-size:12pt;vertical-align:middle;white-space:nowrap;direction:ltr;padding:0px 3px 0px 3px;}
    .ritz .waffle td { border: 1px solid #d0d0d0; }
    /* Auto-wrap for long content */
    .ritz .waffle td.wrap-text {
        white-space: normal !important;
        word-break: break-word !important;
        min-height: 30px;
        height: auto !important;
    }
</style>
<div class="ritz grid-container" dir="ltr">
<table class="waffle" cellspacing="0" cellpadding="0" style="table-layout: fixed; width: 100%;">
    <colgroup>
        <col style="width:223px;"><col style="width:162px;"><col style="width:86px;"><col style="width:100px;"><col style="width:100px;"><col style="width:100px;"><col style="width:81px;"><col style="width:15px;"><col style="width:148px;"><col style="width:213px;"><col style="width:100px;"><col style="width:100px;"><col style="width:100px;"><col style="width:100px;"><col style="width:18px;">
    </colgroup>
    <tbody>
        <tr style="height: 19px;"><td class="s0" dir="ltr" colspan="15">SITE INFORMATION REPORT</td></tr>
        <tr style="height: 19px;"><td class="s1" dir="ltr" colspan="7">General Information</td><td class="s1"></td><td class="s1" dir="ltr" colspan="7">Location</td></tr>
        <tr style="height: 9px;"><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s3"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s3"></td></tr>
        <tr style="height: 19px;"><td class="s2" dir="ltr">Trade Area Name</td><td class="s2"></td><td class="s4" dir="ltr" colspan="5">{{TRADE AREA}}</td><td class="s3"></td><td class="s5" dir="ltr" colspan="2">Site Name</td><td class="s4" dir="ltr" colspan="5">{{SITE NAME}}</td></tr>
        <tr style="height: 19px;"><td class="s2" dir="ltr">Site Name:</td><td class="s2"></td><td class="s4" dir="ltr" colspan="5">{{SITE NAME}}</td><td class="s3"></td><td class="s5" dir="ltr" colspan="2">Unit #, Bldg/St # and St Name</td><td class="s4" dir="ltr" colspan="5">{{UNIT #, BLDG/ST # AND ST NAME}}</td></tr>
        <tr style="height: 19px;"><td class="s2" dir="ltr">Site Number:</td><td class="s2"></td><td class="s4" dir="ltr" colspan="5">{{SITE NO}}</td><td class="s3"></td><td class="s5" dir="ltr" colspan="2">Barangay/District Name</td><td class="s4" dir="ltr" colspan="5">{{BARANGAY/DISTRICT NAME}}</td></tr>
        <tr style="height: 19px;"><td class="s2" dir="ltr">Date Started</td><td class="s2"></td><td class="s4" dir="ltr" colspan="5">{{TIMESTAMP}}</td><td class="s3"></td><td class="s5" dir="ltr" colspan="2">City/Municipality</td><td class="s4" dir="ltr" colspan="5">{{CITY/MUNICIPALITY}}</td></tr>
        <tr style="height: 19px;"><td class="s5" dir="ltr" colspan="2">Date Report Submitted</td><td class="s4" dir="ltr" colspan="5">{{DATE OF REPORT}}</td><td class="s3"></td><td class="s5" dir="ltr" colspan="2">Region</td><td class="s4" dir="ltr" colspan="5">{{REGION}}</td></tr>
        <tr style="height: 19px;"><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s3"></td><td class="s5" dir="ltr" colspan="2">Postal Code</td><td class="s4" dir="ltr" colspan="5">{{POSTAL CODE}}</td></tr>
        <tr style="height: 9px;"><td class="s6"></td><td class="s6"></td><td class="s6"></td><td class="s6"></td><td class="s6"></td><td class="s6"></td><td class="s6"></td><td class="s3"></td><td class="s6"></td><td class="s6"></td><td class="s6"></td><td class="s6"></td><td class="s6"></td><td class="s6"></td><td class="s7"></td></tr>
        <tr style="height: 19px;"><td class="s1" dir="ltr" colspan="7">Terms</td><td class="s3"></td><td class="s1" dir="ltr" colspan="7">Rates</td></tr>
        <tr style="height: 19px;"><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s3"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s3"></td></tr>
        <tr style="height: 19px;"><td class="s2" dir="ltr">Site Availability Date</td><td class="s2"></td><td class="s4" dir="ltr" colspan="5">{{SITE AVAILABILITY DATE}}</td><td class="s3"></td><td class="s8" dir="ltr" colspan="2">Monthly Rental Rate (Php)</td><td class="s4" dir="ltr" colspan="5">{{MONTHLY RENTAL RATE}}</td></tr>
        <tr style="height: 19px;"><td class="s2" dir="ltr">COL Start Date</td><td class="s2"></td><td class="s4" dir="ltr" colspan="5">{{COL START DATE}}</td><td class="s3"></td><td class="s8" dir="ltr" colspan="2">Percentage Rent</td><td class="s4" colspan="5"></td></tr>
        <tr style="height: 19px;"><td class="s2" dir="ltr">COL End Date</td><td class="s2"></td><td class="s4" dir="ltr" colspan="5">{{COL END DATE}}</td><td class="s3"></td><td class="s8" dir="ltr" colspan="2">Minimum Guaranteed Rent</td><td class="s4" colspan="5"></td></tr>
        <tr style="height: 19px;"><td class="s2" dir="ltr">Lease Terms</td><td class="s2"></td><td class="s4" dir="ltr" colspan="5">{{LEASE TERMS}}</td><td class="s3"></td><td class="s8" dir="ltr" colspan="2">Annual Escalation Rate (%)</td><td class="s4" dir="ltr" colspan="5">{{ESCALATION}}</td></tr>
        <tr style="height: 19px;"><td class="s6"></td><td class="s6"></td><td class="s6"></td><td class="s6"></td><td class="s6"></td><td class="s6"></td><td class="s6"></td><td class="s3"></td><td class="s8" dir="ltr" colspan="2">Advance Rental (Php)</td><td class="s4" dir="ltr" colspan="5">{{ADVANCE RENTAL}}</td></tr>
        <tr style="height: 19px;"><td class="s1" dir="ltr" colspan="7">Technical Info</td><td class="s3"></td><td class="s8" dir="ltr" colspan="2">Security Deposit Amount (Php)</td><td class="s4" dir="ltr" colspan="5">{{SECURITY DEPOSIT}}</td></tr>
        <tr style="height: 19px;"><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s3"></td><td class="s8" dir="ltr" colspan="2">CUSA Dues</td><td class="s4" dir="ltr" colspan="5">{{CUSA}}</td></tr>
        <tr style="height: 19px;"><td class="s5" dir="ltr" colspan="2">Lot /Floor Area (in sqm)</td><td class="s4" dir="ltr" colspan="5">{{LOT/FLOOR AREA SQM}}</td><td class="s3"></td><td class="s8" dir="ltr" colspan="2">Estimated Revenue Per Mo.</td><td class="s4" colspan="5"></td></tr>
        <tr style="height: 19px;"><td class="s2" dir="ltr">Frontage (in m)</td><td class="s2"></td><td class="s4" dir="ltr" colspan="5">{{FRONTAGE}}</td><td class="s3"></td><td class="s6"></td><td class="s6"></td><td class="s6"></td><td class="s6"></td><td class="s6"></td><td class="s6"></td><td class="s7"></td></tr>
        <tr style="height: 19px;"><td class="s2" dir="ltr">Depth (in m)</td><td class="s2"></td><td class="s4" colspan="5"></td><td class="s3"></td><td class="s1" dir="ltr" colspan="7">Provisions</td></tr>
        <tr style="height: 19px;"><td class="s5" dir="ltr" colspan="2">Floor to Slab Height (in m) - if Bldg</td><td class="s4" colspan="5"></td><td class="s3"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s3"></td></tr>
        <tr style="height: 19px;"><td class="s5" dir="ltr" colspan="2">No. of Storeys (If Bldg Lessee)</td><td class="s4" colspan="5"></td><td class="s3"></td><td class="s5" dir="ltr" colspan="2">Tenant is the Owner</td><td class="s9" colspan="5"></td></tr>
        <tr style="height: 19px;"><td class="s5" dir="ltr" colspan="2">Type of Structure(if Bldg Lessee)</td><td class="s4" colspan="5"></td><td class="s3"></td><td class="s5" dir="ltr" colspan="2">Lease Type</td><td class="s9" dir="ltr" colspan="5">{{LEASE TYPE}}</td></tr>
        <tr style="height: 19px;"><td class="s2" dir="ltr">Soil Profile</td><td class="s2"></td><td class="s4" colspan="5"></td><td class="s3"></td><td class="s5" dir="ltr" colspan="2">Principal COL</td><td class="s9" colspan="5"></td></tr>
        <tr style="height: 19px;"><td class="s2" dir="ltr">Supply Access:</td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s3"></td><td class="s5" dir="ltr" colspan="2">Sub-Lease Provison</td><td class="s9" colspan="5"></td></tr>
        <tr style="height: 19px;"><td class="s2" dir="ltr">Power</td><td class="s10"></td><td class="s2" dir="ltr">Aircon</td><td class="s10"></td><td class="s5" dir="ltr" colspan="2">LPG Fire Pro</td><td class="s10"></td><td class="s3"></td><td class="s5" dir="ltr" colspan="2">Pre-Term/Partial Term</td><td class="s9" colspan="5"></td></tr>
        <tr style="height: 19px;"><td class="s2" dir="ltr">Water</td><td class="s10"></td><td class="s2" dir="ltr">Exhaust</td><td class="s10"></td><td class="s5" dir="ltr" colspan="2">Drainage TP</td><td class="s10"></td><td class="s3"></td><td class="s5" dir="ltr" colspan="2">Tripartite Agreement</td><td class="s9" colspan="5"></td></tr>
        <tr style="height: 9px;"><td class="s6"></td><td class="s6"></td><td class="s6"></td><td class="s6"></td><td class="s6"></td><td class="s6"></td><td class="s6"></td><td class="s3"></td><td class="s6"></td><td class="s6"></td><td class="s6"></td><td class="s6"></td><td class="s6"></td><td class="s6"></td><td class="s7"></td></tr>
        <tr style="height: 19px;"><td class="s1" dir="ltr" colspan="7">Lessor and Tenant Details</td><td class="s3"></td><td class="s1" dir="ltr" colspan="7">If with Sub-Lessor/ Sub-Lessee</td></tr>
        <tr style="height: 9px;"><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s3"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s3"></td></tr>
        <tr style="height: 19px;"><td class="s2" dir="ltr">Name of Lessor</td><td class="s2"></td><td class="s4" dir="ltr" colspan="5">{{LESSOR}}</td><td class="s3"></td><td class="s5" dir="ltr" colspan="2">Name of Sub-Lessor</td><td class="s9" colspan="5"></td></tr>
        <tr style="height: 19px;"><td class="s2" dir="ltr">Contact No.</td><td class="s2"></td><td class="s4" colspan="5"></td><td class="s3"></td><td class="s5" dir="ltr" colspan="2">Contact No.</td><td class="s9" colspan="5"></td></tr>
        <tr style="height: 19px;"><td class="s2" dir="ltr">E-mail Address</td><td class="s2"></td><td class="s4" colspan="5"></td><td class="s3"></td><td class="s5" dir="ltr" colspan="2">E-mail Address</td><td class="s9" colspan="5"></td></tr>
        <tr style="height: 19px;"><td class="s2" dir="ltr">Type of Ownership</td><td class="s2"></td><td class="s4" colspan="5"></td><td class="s3"></td><td class="s5" dir="ltr" colspan="2">Type of Ownership</td><td class="s9" colspan="5"></td></tr>
        <tr style="height: 19px;"><td class="s2" dir="ltr">Company Name</td><td class="s2"></td><td class="s4" colspan="5"></td><td class="s3"></td><td class="s5" dir="ltr" colspan="2">Company Name</td><td class="s9" colspan="5"></td></tr>
        <tr style="height: 19px;"><td class="s5" dir="ltr" colspan="2">Developer Account Name</td><td class="s4" colspan="5"></td><td class="s3"></td><td class="s5" dir="ltr" colspan="2">Developer Account Name</td><td class="s9" colspan="5"></td></tr>
        <tr style="height: 19px;"><td class="s2" dir="ltr">Business Address</td><td class="s2"></td><td class="s4" colspan="5"></td><td class="s3"></td><td class="s5" dir="ltr" colspan="2">Business Address</td><td class="s9" colspan="5"></td></tr>
        <tr style="height: 19px;"><td class="s5" dir="ltr" colspan="2">Name of Authorized Representative</td><td class="s4" dir="ltr" colspan="5">{{CONTACT PERSON/SOURCE}}</td><td class="s3"></td><td class="s5" dir="ltr" colspan="2">Name of Authorized Representative</td><td class="s9" colspan="5"></td></tr>
        <tr style="height: 19px;"><td class="s5" dir="ltr" colspan="2">Residence Address of Authorized Representative</td><td class="s4" colspan="5"></td><td class="s3"></td><td class="s5" dir="ltr" colspan="2">Residence Address of Authorized Representative</td><td class="s9" colspan="5"></td></tr>
        <tr style="height: 19px;"><td class="s2" dir="ltr">Contact No.</td><td class="s2"></td><td class="s4" dir="ltr" colspan="5">{{CONTACT NUMBER}}</td><td class="s3"></td><td class="s5" dir="ltr" colspan="2">Contact No.</td><td class="s9" colspan="5"></td></tr>
        <tr style="height: 19px;"><td class="s2" dir="ltr">E-mail Address</td><td class="s2"></td><td class="s4" dir="ltr" colspan="5">{{EMAIL ADDRESS}}</td><td class="s3"></td><td class="s5" dir="ltr" colspan="2">E-mail Address</td><td class="s9" colspan="5"></td></tr>
        <tr style="height: 9px;"><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s3"></td><td class="s2"></td><td class="s2"></td><td class="s3" colspan="5"></td></tr>
        <tr style="height: 19px;"><td class="s2" dir="ltr">Name of Lessee</td><td class="s2"></td><td class="s4" colspan="5"></td><td class="s3"></td><td class="s5" dir="ltr" colspan="2">Name of Sub-Lessee</td><td class="s9" colspan="5"></td></tr>
        <tr style="height: 19px;"><td class="s2" dir="ltr">Position</td><td class="s2"></td><td class="s4" colspan="5"></td><td class="s3"></td><td class="s5" dir="ltr" colspan="2">Position</td><td class="s9" colspan="5"></td></tr>
        <tr style="height: 19px;"><td class="s2" dir="ltr">Contact No.</td><td class="s2"></td><td class="s4" colspan="5"></td><td class="s3"></td><td class="s5" dir="ltr" colspan="2">Contact No.</td><td class="s9" colspan="5"></td></tr>
        <tr style="height: 19px;"><td class="s2" dir="ltr">E-mail Address</td><td class="s2"></td><td class="s4" colspan="5"></td><td class="s3"></td><td class="s5" dir="ltr" colspan="2">E-mail Address</td><td class="s9" colspan="5"></td></tr>
        <tr style="height: 19px;"><td class="s5" dir="ltr" colspan="2">Name of Authorized Representative</td><td class="s4" colspan="5"></td><td class="s3"></td><td class="s5" dir="ltr" colspan="2">Name of Authorized Representative</td><td class="s9" colspan="5"></td></tr>
        <tr style="height: 19px;"><td class="s2" dir="ltr">Business Address</td><td class="s2"></td><td class="s4" colspan="5"></td><td class="s3"></td><td class="s5" dir="ltr" colspan="2">Business Address</td><td class="s9" colspan="5"></td></tr>
        <tr style="height: 9px;"><td class="s11"></td><td class="s11"></td><td class="s11"></td><td class="s11"></td><td class="s11"></td><td class="s11"></td><td class="s11"></td><td class="s12"></td><td class="s11"></td><td class="s11"></td><td class="s11"></td><td class="s11"></td><td class="s11"></td><td class="s11"></td><td class="s12"></td></tr>
        <tr style="height: 19px;"><td class="s13" dir="ltr" colspan="15">Regulatory</td></tr>
        <tr style="height: 19px;">
            <td class="s14" dir="ltr">Setback Requirement</td><td class="s15" colspan="4"></td><td class="s16" dir="ltr" colspan="2">Perm Traffic Re-Routing</td><td class="s17"></td><td class="s15" colspan="2"></td><td class="s18" dir="ltr" colspan="5">Future Development</td>
        </tr>
        <tr style="height: 19px;">
            <td class="s14" dir="ltr">Road Widening</td><td class="s15" colspan="4"></td><td class="s16" dir="ltr" colspan="2">Perm Road Closure</td><td class="s17"></td><td class="s15" colspan="2"></td><td class="s18" dir="ltr" colspan="5">Zoning Clearance</td>
        </tr>
        <tr style="height: 19px;">
            <td class="s19" dir="ltr">Pedestrian Overpass</td><td class="s20" colspan="4"></td><td class="s21" dir="ltr" colspan="2">Infrastructure Programs</td><td class="s22"></td><td class="s20" colspan="2"></td><td class="s23" dir="ltr" colspan="5">Gas Station</td>
        </tr>
        <tr style="height: 9px;"><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s3"></td></tr>
        <tr style="height: 19px;"><td class="s24" dir="ltr">Site Acquirability:</td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s3"></td></tr>
        <tr style="height: 19px;"><td class="s2" dir="ltr">Confidence Level</td><td class="s4" colspan="2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s3"></td></tr>
        <tr style="height: 19px;"><td class="s2" dir="ltr">Site Availability</td><td class="s25 softmerge" dir="ltr"><div class="softmerge-inner" style="width:246px;left:-1px">{{SITE AVAILABILITY CLASS}}</div></td><td class="s26" dir="ltr"></td><td class="s27" dir="ltr"></td><td class="s2" dir="ltr"></td><td class="s2" dir="ltr"></td><td class="s2" dir="ltr"></td><td class="s2" dir="ltr"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s2"></td><td class="s3"></td></tr>
        <tr style="height: 19px;"><td class="s6" dir="ltr">Other Remarks:</td><td class="s5" dir="ltr" colspan="7">{{REMARKS}}</td><td class="s6"></td><td class="s6"></td><td class="s6"></td><td class="s6"></td><td class="s6"></td><td class="s6"></td><td class="s7"></td></tr>
    </tbody>
</table>
</div>
"""

# --- PHOTOS HTML TEMPLATE ---
PHOTOS_TEMPLATE_HTML = """
<meta http-equiv="Content-Type" content="text/html; charset=utf-8">
<style type="text/css">
    .ritz .waffle a { color: inherit; }
    .ritz .waffle .s0{background-color:#ffffff;text-align:center;color:#000000;font-family:Arial;font-size:10pt;vertical-align:middle;white-space:nowrap;direction:ltr;padding:0px 3px 0px 3px;}
</style>
<div class="ritz grid-container" dir="ltr">
<table class="waffle" cellspacing="0" cellpadding="0">
    <thead>
        <tr>
            <th class="row-header freezebar-origin-ltr"></th>
            <th id="147604394C0" style="width:100px;" class="column-headers-background">A</th>
            <th id="147604394C1" style="width:100px;" class="column-headers-background">B</th>
            <th id="147604394C2" style="width:100px;" class="column-headers-background">C</th>
            <th id="147604394C3" style="width:100px;" class="column-headers-background">D</th>
            <th id="147604394C4" style="width:100px;" class="column-headers-background">E</th>
            <th id="147604394C5" style="width:100px;" class="column-headers-background">F</th>
            <th id="147604394C6" style="width:100px;" class="column-headers-background">G</th>
            <th id="147604394C7" style="width:27px;" class="column-headers-background">H</th>
            <th id="147604394C8" style="width:100px;" class="column-headers-background">I</th>
            <th id="147604394C9" style="width:100px;" class="column-headers-background">J</th>
            <th id="147604394C10" style="width:100px;" class="column-headers-background">K</th>
            <th id="147604394C11" style="width:100px;" class="column-headers-background">L</th>
            <th id="147604394C12" style="width:100px;" class="column-headers-background">M</th>
            <th id="147604394C13" style="width:100px;" class="column-headers-background">N</th>
            <th id="147604394C14" style="width:100px;" class="column-headers-background">O</th>
        </tr>
    </thead>
    <tbody>
        <tr style="height: 19px">
            <th id="147604394R0" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">1</div></th>
            <td></td><td></td><td></td><td></td><td></td><td></td><td></td>
            <td></td>
            <td></td><td></td><td></td><td></td><td></td><td></td><td></td>
        </tr>
        <tr style="height: 19px">
            <th id="147604394R1" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">2</div></th>
            <td class="s0" dir="ltr" colspan="7" rowspan="18">{{PROPERTY PHOTO 1}}</td>
            <td></td>
            <td class="s0" dir="ltr" colspan="7" rowspan="18">{{PROPERTY PHOTO 2}}</td>
        </tr>
        <tr style="height: 19px"><th id="147604394R2" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">3</div></th><td></td></tr>
        <tr style="height: 19px"><th id="147604394R3" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">4</div></th><td></td></tr>
        <tr style="height: 19px"><th id="147604394R4" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">5</div></th><td></td></tr>
        <tr style="height: 19px"><th id="147604394R5" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">6</div></th><td></td></tr>
        <tr style="height: 19px"><th id="147604394R6" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">7</div></th><td></td></tr>
        <tr style="height: 19px"><th id="147604394R7" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">8</div></th><td></td></tr>
        <tr style="height: 19px"><th id="147604394R8" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">9</div></th><td></td></tr>
        <tr style="height: 19px"><th id="147604394R9" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">10</div></th><td></td></tr>
        <tr style="height: 19px"><th id="147604394R10" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">11</div></th><td></td></tr>
        <tr style="height: 19px"><th id="147604394R11" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">12</div></th><td></td></tr>
        <tr style="height: 19px"><th id="147604394R12" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">13</div></th><td></td></tr>
        <tr style="height: 19px"><th id="147604394R13" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">14</div></th><td></td></tr>
        <tr style="height: 19px"><th id="147604394R14" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">15</div></th><td></td></tr>
        <tr style="height: 19px"><th id="147604394R15" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">16</div></th><td></td></tr>
        <tr style="height: 19px"><th id="147604394R16" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">17</div></th><td></td></tr>
        <tr style="height: 19px"><th id="147604394R17" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">18</div></th><td></td></tr>
        <tr style="height: 19px"><th id="147604394R18" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">19</div></th><td></td></tr>
        <tr style="height: 19px">
            <th id="147604394R19" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">20</div></th>
            <td></td><td></td><td></td><td></td><td></td><td></td><td></td>
            <td></td>
            <td></td><td></td><td></td><td></td><td></td><td></td><td></td>
        </tr>
        <tr style="height: 19px">
            <th id="147604394R20" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">21</div></th>
            <td class="s0" dir="ltr" colspan="7" rowspan="18">{{PROPERTY PHOTO 3}}</td>
            <td></td>
            <td class="s0" dir="ltr" colspan="7" rowspan="18">{{PROPERTY PHOTO 4}}</td>
        </tr>
        <tr style="height: 19px"><th id="147604394R21" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">22</div></th><td></td></tr>
        <tr style="height: 19px"><th id="147604394R22" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">23</div></th><td></td></tr>
        <tr style="height: 19px"><th id="147604394R23" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">24</div></th><td></td></tr>
        <tr style="height: 19px"><th id="147604394R24" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">25</div></th><td></td></tr>
        <tr style="height: 19px"><th id="147604394R25" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">26</div></th><td></td></tr>
        <tr style="height: 19px"><th id="147604394R26" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">27</div></th><td></td></tr>
        <tr style="height: 19px"><th id="147604394R27" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">28</div></th><td></td></tr>
        <tr style="height: 19px"><th id="147604394R28" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">29</div></th><td></td></tr>
        <tr style="height: 19px"><th id="147604394R29" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">30</div></th><td></td></tr>
        <tr style="height: 19px"><th id="147604394R30" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">31</div></th><td></td></tr>
        <tr style="height: 19px"><th id="147604394R31" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">32</div></th><td></td></tr>
        <tr style="height: 19px"><th id="147604394R32" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">33</div></th><td></td></tr>
        <tr style="height: 19px"><th id="147604394R33" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">34</div></th><td></td></tr>
        <tr style="height: 19px"><th id="147604394R34" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">35</div></th><td></td></tr>
        <tr style="height: 19px"><th id="147604394R35" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">36</div></th><td></td></tr>
        <tr style="height: 19px"><th id="147604394R36" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">37</div></th><td></td></tr>
        <tr style="height: 19px"><th id="147604394R37" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">38</div></th><td></td></tr>
        <tr style="height: 19px">
            <th id="147604394R38" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">39</div></th>
            <td></td><td></td><td></td><td></td><td></td><td></td><td></td>
            <td></td>
            <td></td><td></td><td></td><td></td><td></td><td></td><td></td>
        </tr>
        <tr style="height: 19px">
            <th id="147604394R39" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">40</div></th>
            <td class="s0" dir="ltr" colspan="7" rowspan="18">{{PROPERTY PHOTO 5}}</td>
            <td></td>
            <td></td><td></td><td></td><td></td><td></td><td></td><td></td>
        </tr>
        <tr style="height: 19px"><th id="147604394R40" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">41</div></th><td></td><td></td><td></td><td></td><td></td><td></td><td></td><td></td></tr>
        <tr style="height: 19px"><th id="147604394R41" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">42</div></th><td></td><td></td><td></td><td></td><td></td><td></td><td></td><td></td></tr>
        <tr style="height: 19px"><th id="147604394R42" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">43</div></th><td></td><td></td><td></td><td></td><td></td><td></td><td></td><td></td></tr>
        <tr style="height: 19px"><th id="147604394R43" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">44</div></th><td></td><td></td><td></td><td></td><td></td><td></td><td></td><td></td></tr>
        <tr style="height: 19px"><th id="147604394R44" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">45</div></th><td></td><td></td><td></td><td></td><td></td><td></td><td></td><td></td></tr>
        <tr style="height: 19px"><th id="147604394R45" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">46</div></th><td></td><td></td><td></td><td></td><td></td><td></td><td></td><td></td></tr>
        <tr style="height: 19px"><th id="147604394R46" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">47</div></th><td></td><td></td><td></td><td></td><td></td><td></td><td></td><td></td></tr>
        <tr style="height: 19px"><th id="147604394R47" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">48</div></th><td></td><td></td><td></td><td></td><td></td><td></td><td></td><td></td></tr>
        <tr style="height: 19px"><th id="147604394R48" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">49</div></th><td></td><td></td><td></td><td></td><td></td><td></td><td></td><td></td></tr>
        <tr style="height: 19px"><th id="147604394R49" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">50</div></th><td></td><td></td><td></td><td></td><td></td><td></td><td></td><td></td></tr>
        <tr style="height: 19px"><th id="147604394R50" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">51</div></th><td></td><td></td><td></td><td></td><td></td><td></td><td></td><td></td></tr>
        <tr style="height: 19px"><th id="147604394R51" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">52</div></th><td></td><td></td><td></td><td></td><td></td><td></td><td></td><td></td></tr>
        <tr style="height: 19px"><th id="147604394R52" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">53</div></th><td></td><td></td><td></td><td></td><td></td><td></td><td></td><td></td></tr>
        <tr style="height: 19px"><th id="147604394R53" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">54</div></th><td></td><td></td><td></td><td></td><td></td><td></td><td></td><td></td></tr>
        <tr style="height: 19px"><th id="147604394R54" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">55</div></th><td></td><td></td><td></td><td></td><td></td><td></td><td></td><td></td></tr>
        <tr style="height: 19px"><th id="147604394R55" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">56</div></th><td></td><td></td><td></td><td></td><td></td><td></td><td></td><td></td></tr>
        <tr style="height: 19px"><th id="147604394R56" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">57</div></th><td></td><td></td><td></td><td></td><td></td><td></td><td></td><td></td></tr>
    </tbody>
</table>
</div>
"""

# --- DOCS HTML TEMPLATE ---
DOCS_TEMPLATE_HTML = """
<meta http-equiv="Content-Type" content="text/html; charset=utf-8">
<style type="text/css">
    .ritz .waffle a { color: inherit; }
    .ritz .waffle .s0{background-color:#ffffff;text-align:center;color:#000000;font-family:Arial;font-size:10pt;vertical-align:middle;white-space:nowrap;direction:ltr;padding:0px 3px 0px 3px;}
</style>
<div class="ritz grid-container" dir="ltr">
<table class="waffle" cellspacing="0" cellpadding="0">
    <thead>
        <tr>
            <th class="row-header freezebar-origin-ltr"></th>
            <th id="1206708472C0" style="width:100px;" class="column-headers-background">A</th>
            <th id="1206708472C1" style="width:100px;" class="column-headers-background">B</th>
            <th id="1206708472C2" style="width:100px;" class="column-headers-background">C</th>
            <th id="1206708472C3" style="width:100px;" class="column-headers-background">D</th>
            <th id="1206708472C4" style="width:100px;" class="column-headers-background">E</th>
            <th id="1206708472C5" style="width:100px;" class="column-headers-background">F</th>
            <th id="1206708472C6" style="width:100px;" class="column-headers-background">G</th>
            <th id="1206708472C7" style="width:100px;" class="column-headers-background">H</th>
            <th id="1206708472C8" style="width:100px;" class="column-headers-background">I</th>
            <th id="1206708472C9" style="width:100px;" class="column-headers-background">J</th>
            <th id="1206708472C10" style="width:100px;" class="column-headers-background">K</th>
        </tr>
    </thead>
    <tbody>
        <tr style="height: 19px">
            <th id="1206708472R0" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">1</div></th>
            <td class="s0" dir="ltr" colspan="5" rowspan="33">{{TCT}}</td>
            <td></td>
            <td class="s0" dir="ltr" colspan="5" rowspan="33">{{TAX MAP}}</td>
        </tr>
        <tr style="height: 19px"><th id="1206708472R1" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">2</div></th><td></td></tr>
        <tr style="height: 19px"><th id="1206708472R2" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">3</div></th><td></td></tr>
        <tr style="height: 19px"><th id="1206708472R3" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">4</div></th><td></td></tr>
        <tr style="height: 19px"><th id="1206708472R4" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">5</div></th><td></td></tr>
        <tr style="height: 19px"><th id="1206708472R5" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">6</div></th><td></td></tr>
        <tr style="height: 19px"><th id="1206708472R6" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">7</div></th><td></td></tr>
        <tr style="height: 19px"><th id="1206708472R7" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">8</div></th><td></td></tr>
        <tr style="height: 19px"><th id="1206708472R8" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">9</div></th><td></td></tr>
        <tr style="height: 19px"><th id="1206708472R9" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">10</div></th><td></td></tr>
        <tr style="height: 19px"><th id="1206708472R10" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">11</div></th><td></td></tr>
        <tr style="height: 19px"><th id="1206708472R11" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">12</div></th><td></td></tr>
        <tr style="height: 19px"><th id="1206708472R12" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">13</div></th><td></td></tr>
        <tr style="height: 19px"><th id="1206708472R13" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">14</div></th><td></td></tr>
        <tr style="height: 19px"><th id="1206708472R14" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">15</div></th><td></td></tr>
        <tr style="height: 19px"><th id="1206708472R15" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">16</div></th><td></td></tr>
        <tr style="height: 19px"><th id="1206708472R16" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">17</div></th><td></td></tr>
        <tr style="height: 19px"><th id="1206708472R17" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">18</div></th><td></td></tr>
        <tr style="height: 19px"><th id="1206708472R18" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">19</div></th><td></td></tr>
        <tr style="height: 19px"><th id="1206708472R19" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">20</div></th><td></td></tr>
        <tr style="height: 19px"><th id="1206708472R20" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">21</div></th><td></td></tr>
        <tr style="height: 19px"><th id="1206708472R21" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">22</div></th><td></td></tr>
        <tr style="height: 19px"><th id="1206708472R22" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">23</div></th><td></td></tr>
        <tr style="height: 19px"><th id="1206708472R23" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">24</div></th><td></td></tr>
        <tr style="height: 19px"><th id="1206708472R24" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">25</div></th><td></td></tr>
        <tr style="height: 19px"><th id="1206708472R25" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">26</div></th><td></td></tr>
        <tr style="height: 19px"><th id="1206708472R26" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">27</div></th><td></td></tr>
        <tr style="height: 19px"><th id="1206708472R27" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">28</div></th><td></td></tr>
        <tr style="height: 19px"><th id="1206708472R28" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">29</div></th><td></td></tr>
        <tr style="height: 19px"><th id="1206708472R29" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">30</div></th><td></td></tr>
        <tr style="height: 19px"><th id="1206708472R30" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">31</div></th><td></td></tr>
        <tr style="height: 19px"><th id="1206708472R31" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">32</div></th><td></td></tr>
        <tr style="height: 19px"><th id="1206708472R32" style="height: 19px;" class="row-headers-background"><div class="row-header-wrapper" style="line-height: 19px">33</div></th><td></td></tr>
    </tbody>
</table>
</div>
"""

# --- LOAD DATA ---
@st.cache_data(ttl=3600)
def load_data():
    source_data = download_file(SOURCE_URL)
    template_data = download_file(TEMPLATE_URL)
    if source_data is None or template_data is None: return None, None, None
    
    df = pd.read_excel(source_data)
    df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
    df.columns = df.columns.str.strip().str.upper()
    
    temp_wb = load_workbook(template_data)
    placeholders = get_placeholders(temp_wb.active)
    template_data.seek(0)
    return df, placeholders, template_data.getvalue()

with st.spinner("Loading Data Assets..."):
    df, placeholders, template_bytes_raw = load_data()

if df is None or template_bytes_raw is None:
    st.error("Failed to load data. Please check connection profiles.")
    st.stop()

template_data = io.BytesIO(template_bytes_raw)

# --- CONTROLS ROW ---
# Create combined display for dropdown: [SITE NO] - [SITE NAME]
df['SITE DISPLAY'] = df.apply(
    lambda row: f"{row.get('SITE NO', '')} - {row.get('SITE NAME', '')}" if pd.notna(row.get('SITE NO')) else row.get('SITE NAME', ''),
    axis=1
)

trade_areas = sorted(df["TRADE AREA"].dropna().unique())
col1, col2, col3, col4 = st.columns([2.0, 2.0, 0.8, 1.2])

with col1:
    selected_ta = st.selectbox("Select Trade Area", options=trade_areas, index=0 if trade_areas else None, key="ta_select")

with col2:
    if selected_ta:
        sites_in_ta = df[df["TRADE AREA"] == selected_ta].sort_values(['SITE NO', 'SITE NAME'])
        site_options = sites_in_ta['SITE DISPLAY'].tolist()
        site_values = sites_in_ta['SITE NAME'].tolist()
        selected_display = st.selectbox("Select Site Name", options=site_options, index=0 if site_options else None, key="site_select")
        selected_site = site_values[site_options.index(selected_display)] if selected_display and selected_display in site_options else None
    else:
        selected_site = None
        st.selectbox("Select Site Name", options=[], disabled=True, key="site_select")

with col3:
    if st.button("Refresh", use_container_width=True):
        st.cache_data.clear()
        st.rerun()

site_excel_bytes = None
site_row_data = None

if selected_ta and selected_site:
    site_data = df[(df["TRADE AREA"] == selected_ta) & (df["SITE NAME"] == selected_site)]
    if not site_data.empty:
        site_row_data = site_data.iloc[0]
        template_data.seek(0)
        wb = load_workbook(template_data)
        base_sheet = wb.active
        for row_cells in base_sheet.iter_rows():
            for cell in row_cells:
                if isinstance(cell.value, str) and ("{{" in cell.value):
                    new_val = cell.value
                    for ph in placeholders:
                        target_regex = r"\{\{\s*" + re.escape(ph) + r"(\s*:.*?)?\}\}"
                        if re.search(target_regex, new_val):
                            raw_data_val = site_row_data.get(ph.upper(), "")
                            if pd.isna(raw_data_val) or raw_data_val is None:
                                val_str = ""  # Empty for no match
                            else:
                                if isinstance(raw_data_val, float) and raw_data_val.is_integer():
                                    val_str = str(int(raw_data_val))
                                elif hasattr(raw_data_val, 'strftime'):
                                    val_str = raw_data_val.strftime('%B %d, %Y')
                                else:
                                    val_str = str(raw_data_val)
                            new_val = re.sub(target_regex, val_str, new_val)
                    cell.value = new_val.strip() if new_val else ""
        
        # Set DATE OF REPORT to today's date only in the Excel export
        for row_cells in base_sheet.iter_rows():
            for cell in row_cells:
                if isinstance(cell.value, str) and "{{DATE OF REPORT}}" in cell.value:
                    cell.value = get_today_date()
        
        ex_buf = io.BytesIO()
        wb.save(ex_buf)
        site_excel_bytes = ex_buf.getvalue()

with col4:
    if selected_ta:
        with io.BytesIO() as wb_buffer:
            template_data.seek(0)
            wb_bulk = load_workbook(template_data)
            base_sheet_bulk = wb_bulk.active
            base_sheet_bulk.title = "TEMPLATE_TO_DELETE"
            existing_tabs_bulk = set()
            
            ta_rows = df[df["TRADE AREA"] == selected_ta].sort_values(['SITE NO', 'SITE NAME'])
            for _, r_row in ta_rows.iterrows():
                s_name = r_row.get("SITE NAME", "Unknown")
                safe_tab_name = sanitize_tab_name(s_name, existing_tabs_bulk)
                new_sheet = wb_bulk.copy_worksheet(base_sheet_bulk)
                new_sheet.title = safe_tab_name
                for row_cells in new_sheet.iter_rows():
                    for cell in row_cells:
                        if isinstance(cell.value, str) and ("{{" in cell.value):
                            new_val = cell.value
                            for ph in placeholders:
                                target_regex = r"\{\{\s*" + re.escape(ph) + r"(\s*:.*?)?\}\}"
                                if re.search(target_regex, new_val):
                                    raw_data_val = r_row.get(ph.upper(), "")
                                    if pd.isna(raw_data_val) or raw_data_val is None:
                                        val_str = ""  # Empty for no match
                                    else:
                                        if isinstance(raw_data_val, float) and raw_data_val.is_integer():
                                            val_str = str(int(raw_data_val))
                                        elif hasattr(raw_data_val, 'strftime'):
                                            val_str = raw_data_val.strftime('%B %d, %Y')
                                        else:
                                            val_str = str(raw_data_val)
                                    new_val = re.sub(target_regex, val_str, new_val)
                            cell.value = new_val.strip() if new_val else ""
                
                # Set DATE OF REPORT to today's date for each sheet in export
                for row_cells in new_sheet.iter_rows():
                    for cell in row_cells:
                        if isinstance(cell.value, str) and "{{DATE OF REPORT}}" in cell.value:
                            cell.value = get_today_date()
            
            wb_bulk.remove(base_sheet_bulk)
            wb_bulk.save(wb_buffer)
            
            st.download_button(
                "Export", 
                data=wb_buffer.getvalue(), 
                file_name=f"Trade_Area_Report_{str(selected_ta).replace('/', '-')}.xlsx", 
                use_container_width=True
            )

# --- DYNAMIC MULTI-VIEW INTERFACE ROUTER ---
if site_excel_bytes and site_row_data is not None:
    tab1, tab2, tab3 = st.tabs(["PROPERTY DETAILS", "PROPERTY PHOTOS", "PROPERTY DOCS"])
    
    # --- TAB 1: PROPERTY DETAILS ---
    with tab1:
        try:
            def process_val(key_string):
                val = site_row_data.get(key_string.upper(), "")
                if pd.isna(val) or val is None: return ""
                if isinstance(val, float) and val.is_integer(): return str(int(val))
                if hasattr(val, 'strftime'): return val.strftime('%B %d, %Y')
                return str(val).strip()

            rendered_view = RAW_TEMPLATE_HTML
            
            # Replace all placeholders with actual values (or empty string if no match)
            placeholder_map = {
                "{{TRADE AREA}}": process_val("TRADE AREA"),
                "{{SITE NAME}}": process_val("SITE NAME"),
                "{{SITE NO}}": process_val("SITE NO"),
                "{{TIMESTAMP}}": process_val("TIMESTAMP"),
                "{{UNIT #, BLDG/ST # AND ST NAME}}": process_val("UNIT #, BLDG/ST # AND ST NAME"),
                "{{BARANGAY/DISTRICT NAME}}": process_val("BARANGAY/DISTRICT NAME"),
                "{{CITY/MUNICIPALITY}}": process_val("CITY/MUNICIPALITY"),
                "{{REGION}}": process_val("REGION"),
                "{{POSTAL CODE}}": process_val("POSTAL CODE"),
                "{{MONTHLY RENTAL RATE}}": process_val("MONTHLY RENTAL RATE"),
                "{{ESCALATION}}": process_val("ESCALATION"),
                "{{ADVANCE RENTAL}}": process_val("ADVANCE RENTAL"),
                "{{SECURITY DEPOSIT}}": process_val("SECURITY DEPOSIT"),
                "{{CUSA}}": process_val("CUSA"),
                "{{LOT/FLOOR AREA SQM}}": process_val("LOT/FLOOR AREA SQM"),
                "{{LEASE TYPE}}": process_val("LEASE TYPE"),
                "{{LESSOR}}": process_val("LESSOR"),
                "{{CONTACT PERSON/SOURCE}}": process_val("CONTACT PERSON/SOURCE"),
                "{{CONTACT NUMBER}}": process_val("CONTACT NUMBER"),
                "{{EMAIL ADDRESS}}": process_val("EMAIL ADDRESS"),
                "{{SITE AVAILABILITY CLASS}}": process_val("SITE AVAILABILITY CLASS"),
                "{{REMARKS}}": process_val("REMARKS"),
                "{{DATE OF REPORT}}": get_today_date()
            }
            
            for placeholder, value in placeholder_map.items():
                rendered_view = rendered_view.replace(placeholder, value)

            st.markdown(f'<div class="excel-container">{rendered_view}</div>', unsafe_allow_html=True)
        except Exception as e:
            st.error(f"Error rendering details tab blueprint: {str(e)}")

    # --- TAB 2: PROPERTY PHOTOS ---
    with tab2:
        st.markdown(f"### Photos for {selected_site}")
        
        # Get photo URLs from the data - try multiple possible column names
        photo_column_names = ["PHOTOS", "PHOTO", "PHOTO LINKS", "IMAGE", "IMAGES", "PICTURES"]
        raw_photos = ""
        
        for col_name in photo_column_names:
            if col_name in site_row_data.index:
                raw_photos = site_row_data.get(col_name, "")
                if raw_photos and str(raw_photos).strip():
                    break
        
        # Parse photo links
        photo_urls = parse_link_cell(raw_photos)
        
        # Map photo URLs to the placeholders in the template
        # The template has placeholders: {{PROPERTY PHOTO 1}} through {{PROPERTY PHOTO 5}}
        photo_placeholders = {
            "{{PROPERTY PHOTO 1}}": photo_urls[0] if len(photo_urls) > 0 else "",
            "{{PROPERTY PHOTO 2}}": photo_urls[1] if len(photo_urls) > 1 else "",
            "{{PROPERTY PHOTO 3}}": photo_urls[2] if len(photo_urls) > 2 else "",
            "{{PROPERTY PHOTO 4}}": photo_urls[3] if len(photo_urls) > 3 else "",
            "{{PROPERTY PHOTO 5}}": photo_urls[4] if len(photo_urls) > 4 else "",
        }
        
        # Build the photos HTML with actual image links
        photos_html = PHOTOS_TEMPLATE_HTML
        for placeholder, url in photo_placeholders.items():
            if url and str(url).strip():
                # Convert to direct download URL for embedding
                direct_url = transform_to_direct_download(url)
                # Create an image tag that will display in the cell
                img_tag = f'<img src="{direct_url}" style="max-width:100%; max-height:300px; object-fit:contain; border-radius:2px;" onerror="this.style.display=\'none\';">'
                photos_html = photos_html.replace(placeholder, img_tag)
            else:
                photos_html = photos_html.replace(placeholder, "")
        
        st.markdown(f'<div class="excel-container">{photos_html}</div>', unsafe_allow_html=True)

    # --- TAB 3: PROPERTY DOCS ---
    with tab3:
        st.markdown(f"### Documents for {selected_site}")
        
        # Get document URLs from the data - try multiple possible column names
        doc_column_names = ["DOCS", "DOC", "DOCUMENTS", "DOCUMENT LINKS", "FILES", "ATTACHMENTS"]
        raw_docs = ""
        
        for col_name in doc_column_names:
            if col_name in site_row_data.index:
                raw_docs = site_row_data.get(col_name, "")
                if raw_docs and str(raw_docs).strip():
                    break
        
        # Parse document links
        doc_urls = parse_link_cell(raw_docs)
        
        # Map document URLs to the placeholders in the template
        # The template has placeholders: {{TCT}} and {{TAX MAP}}
        doc_placeholders = {
            "{{TCT}}": doc_urls[0] if len(doc_urls) > 0 else "",
            "{{TAX MAP}}": doc_urls[1] if len(doc_urls) > 1 else "",
        }
        
        # Build the documents HTML with actual document links
        docs_html = DOCS_TEMPLATE_HTML
        for placeholder, url in doc_placeholders.items():
            if url and str(url).strip():
                # Create a clickable link that opens in a new tab
                direct_url = transform_to_direct_download(url)
                link_tag = f'<a href="{direct_url}" target="_blank" style="color:#0066cc; text-decoration:underline;">📄 View Document</a>'
                docs_html = docs_html.replace(placeholder, link_tag)
            else:
                docs_html = docs_html.replace(placeholder, "")
        
        st.markdown(f'<div class="excel-container">{docs_html}</div>', unsafe_allow_html=True)

else:
    st.info("Select a Trade Area and Site to view the report.")
