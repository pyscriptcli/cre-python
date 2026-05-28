import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, PatternFill
import re
import io
import zipfile
import difflib
import math
import streamlit as st
import json
import os


st.set_page_config(page_title="Central Portal", layout="centered")

# Ensure this exact block starts and ends with triple double-quotes
css_overrides = """
<style>
#MainMenu {visibility: hidden;}
footer {visibility: hidden;}
header {visibility: hidden;}

html, body, [class*="css"] {
    font-family: Arial, Helvetica, sans-serif !important;
    color: #003366 !important;
}
h1, h2, h3, h4, h5, h6, p, span, label {
    color: #003366 !important;
    font-family: Arial, Helvetica, sans-serif !important;
}

.gold-divider {
    border-top: 0.75px solid #C9AB4C;
    margin: 25px 0;
}

.tool-box {
    background-color: #E8D494;
    padding: 15px;
    margin-bottom: 12px;
    text-align: center;
    display: block;
    text-decoration: none !important;
    color: #003366 !important;
    font-family: Arial, Helvetica, sans-serif;
    font-weight: bold;
    border-radius: 2px;
}
.tool-box:hover {
    background-color: #D6C282;
}
</style>
"""

st.markdown(css_overrides, unsafe_allow_html=True)

tool_name = st.text_input("Tool Name")
tool_url = st.text_input("URL")

if st.button("Add"):
    if tool_name and tool_url:
        data = []
        if os.path.exists("prime_tools.json"):
            with open("prime_tools.json", "r") as f:
                try:
                    data = json.load(f)
                except json.JSONDecodeError:
                    data = []
        
        data.append({"name": tool_name, "url": tool_url})
        
        with open("prime_tools.json", "w") as f:
            json.dump(data, f)
        st.rerun()

st.markdown('<div class="gold-divider"></div>', unsafe_allow_html=True)

if os.path.exists("prime_tools.json"):
    with open("prime_tools.json", "r") as f:
        try:
            tools = json.load(f)
            for tool in tools:
                st.markdown(
                    f'<a href="{tool["url"]}" target="_self" class="tool-box">{tool["name"]}</a>',
                    unsafe_allow_html=True
                )
        except json.JSONDecodeError:
            pass


# Must be the first Streamlit command
st.set_page_config(page_title="Report Generator", layout="centered")

def get_placeholders(sheet):
    """Extract all {{Placeholder}} variables from the Excel sheet."""
    placeholders = set()
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                matches = re.findall(r"\{\{(.*?)\}\}", cell.value)
                placeholders.update(matches)
    return sorted(list(placeholders))

def sanitize_tab_name(name, existing_names):
    """Strip illegal Excel characters, slice to 31 chars, and handle duplicates."""
    illegal_chars = r'[\\/*?\[\]:]'
    clean_name = re.sub(illegal_chars, '', str(name))
    base_name = clean_name[:31]
    
    if base_name not in existing_names:
        existing_names.add(base_name)
        return base_name
    
    counter = 1
    while True:
        suffix = f" ({counter})"
        max_len = 31 - len(suffix)
        new_name = f"{clean_name[:max_len]}{suffix}"
        if new_name not in existing_names:
            existing_names.add(new_name)
            return new_name
        counter += 1

# Reset session state for a fresh run if needed
if "zip_data" not in st.session_state:
    st.session_state.zip_data = None

st.title("Site Information Report")

# STEP 1: The Upload Section
raw_file = st.file_uploader("Upload Raw Data (Excel)", type=["xlsx", "xls"])
template_file = st.file_uploader("Upload Excel Template", type=["xlsx"])

if raw_file and template_file:
    # Read headers from Raw Data & Auto-clean 'Unnamed' blank columns
    df = pd.read_excel(raw_file)
    df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
    df.columns = df.columns.str.strip().str.upper()
    headers = list(df.columns)
    
    # Mandatory Column Check (Moved up to prevent errors early)
    if "TRADE AREA" not in df.columns or "SITE NAME" not in df.columns:
        st.error("ERROR: Raw data must contain exactly 'TRADE AREA' and 'SITE NAME' columns.")
        st.stop()
    
    # Read placeholders from Template
    template_wb = openpyxl.load_workbook(template_file)
    template_sheet = template_wb.active
    placeholders = get_placeholders(template_sheet)
    
    if not placeholders:
        st.warning("No {{Placeholders}} found in the uploaded template.")
    else:
        # STEP 2: The Smart Mapping Section
        st.markdown("### Data Mapping")
        st.markdown("Verify the auto-matched data columns for your template placeholders.")
        
        mapping = {}
        for ph in placeholders:
            match = difflib.get_close_matches(ph, headers, n=1, cutoff=0.3)
            default_index = headers.index(match[0]) if match else 0
            
            col1, col2 = st.columns([1, 2])
            with col1:
                st.markdown(f"**{{{{{ph}}}}}**")
            with col2:
                if ph in ["STREET", "BARANGAY", "CITY", "REGION", "POSTAL"] and "LOCATION/ADDRESS" in headers:
                    default_index = headers.index("LOCATION/ADDRESS")
                
                mapping[ph] = st.selectbox("Map to column:", headers, index=default_index, key=f"map_{ph}", label_visibility="collapsed")
        
        st.divider()

        # STEP 2.5: Trade Area Selection
        st.markdown("### Select Trade Areas")
        st.markdown("Choose which Trade Areas to include in the batch generation.")
        
        # Get unique Trade Areas and cast to string for consistency
        unique_tas = sorted([str(ta) for ta in df["TRADE AREA"].dropna().unique()])
        
        # Select All / Clear All Buttons
        col_sel, col_clr, _ = st.columns([1, 1, 3])
        if col_sel.button("Select All", use_container_width=True):
            for ta in unique_tas:
                st.session_state[f"chk_{ta}"] = True
            st.rerun()
            
        if col_clr.button("Clear All", use_container_width=True):
            for ta in unique_tas:
                st.session_state[f"chk_{ta}"] = False
            st.rerun()

        # Render Checkboxes in a scrollable container
        selected_tas = []
        with st.container(height=250, border=True):
            for ta in unique_tas:
                # Initialize state to True by default if not set
                if f"chk_{ta}" not in st.session_state:
                    st.session_state[f"chk_{ta}"] = True
                
                # Render checkbox tied to session state
                if st.checkbox(ta, key=f"chk_{ta}"):
                    selected_tas.append(ta)
        
        st.divider()

        action_placeholder = st.empty()
        
        if st.session_state.zip_data is None:
            if action_placeholder.button("Generate Reports", use_container_width=True):
                
                if not selected_tas:
                    st.warning("Please select at least one Trade Area to generate reports.")
                else:
                    # STEP 3: The Engine
                    action_placeholder.empty() 
                    progress_bar = st.progress(0)
                    status_text = st.empty()
                    
                    # Filter the dataframe to ONLY include selected Trade Areas
                    filtered_df = df[df["TRADE AREA"].astype(str).isin(selected_tas)]
                    groups = filtered_df.groupby("TRADE AREA")
                    total_groups = len(groups)
                    
                    zip_buffer = io.BytesIO()
                    
                    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
                        for i, (trade_area, group) in enumerate(groups):
                            status_text.text(f"Processing Trade Area: {trade_area}...")
                            
                            template_file.seek(0)
                            wb = openpyxl.load_workbook(template_file)
                            base_sheet = wb.active
                            base_sheet.title = "TEMPLATE_TO_DELETE"
                            existing_tabs = set()
                            
                            for _, row in group.iterrows():
                                site_id = row.get("SITE ID", "NO-ID")
                                site_name = row.get("SITE NAME", site_id)
                                safe_tab_name = sanitize_tab_name(site_name, existing_tabs)
                                
                                new_sheet = wb.copy_worksheet(base_sheet)
                                new_sheet.title = safe_tab_name
                                
                                # Inject Mapped Data & Format
                                for row_cells in new_sheet.iter_rows():
                                    for cell in row_cells:
                                        if isinstance(cell.value, str) and "{{" in cell.value:
                                            original_val = cell.value
                                            new_val = cell.value
                                            injected_length = 0
                                            has_injected_data = False
                                            
                                            for ph, header in mapping.items():
                                                target = f"{{{{{ph}}}}}"
                                                if target in new_val:
                                                    val = row.get(header)
                                                    
                                                    # 1. Date Formatting Rule
                                                    if ("DATE" in str(header).upper() or isinstance(val, pd.Timestamp)) and not pd.isna(val):
                                                        try:
                                                            val_str = pd.to_datetime(val).strftime("%B %d, %Y")
                                                        except:
                                                            val_str = str(val)
                                                            
                                                    # 2. Address Slicing Rule
                                                    elif ph in ["STREET", "BARANGAY", "CITY", "REGION", "POSTAL"] and isinstance(val, str):
                                                        p = [part.strip() for part in val.split(",")]
                                                        length = len(p)
                                                        if ph == "STREET":
                                                            val_str = ", ".join(p[:max(0, length - 6)]) if length >= 6 else val
                                                        elif ph == "BARANGAY":
                                                            val_str = p[length - 6] if length >= 6 else ""
                                                        elif ph == "CITY":
                                                            val_str = p[length - 5] if length >= 5 else ""
                                                        elif ph == "REGION":
                                                            val_str = p[length - 3] if length >= 3 else ""
                                                        elif ph == "POSTAL":
                                                            val_str = p[length - 2] if length >= 2 else ""
                                                    
                                                    # 3. Standard Text Rule
                                                    else:
                                                        val_str = "" if pd.isna(val) else str(val)
                                                    
                                                    # Flag for conditional formatting
                                                    if val_str.strip() != "":
                                                        has_injected_data = True
                                                        injected_length = len(val_str)
                                                    
                                                    new_val = new_val.replace(target, val_str)
                                            
                                            cell.value = new_val
                                            
                                            # --- CONDITIONAL WHITE BACKGROUND FIX ---
                                            if has_injected_data:
                                                cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                                            
                                            # --- THE FORMATTING ALIGNMENT FIX ---
                                            al = cell.alignment
                                            force_left = "DATE" in original_val.upper()
                                            
                                            if al:
                                                cell.alignment = Alignment(
                                                    horizontal='left' if force_left else al.horizontal,
                                                    vertical=al.vertical,
                                                    text_rotation=al.text_rotation,
                                                    wrap_text=True,
                                                    shrink_to_fit=al.shrink_to_fit,
                                                    indent=al.indent
                                                )
                                            else:
                                                cell.alignment = Alignment(horizontal='left' if force_left else 'general', wrap_text=True)
                                                
                                            # --- MERGED CELL HEIGHT OVERRIDE ---
                                            if injected_length > 60:
                                                chars_per_line = 85
                                                estimated_lines = math.ceil(injected_length / chars_per_line)
                                                calculated_height = estimated_lines * 15 
                                                
                                                current_height = new_sheet.row_dimensions[cell.row].height
                                                if current_height is None or calculated_height > current_height:
                                                    new_sheet.row_dimensions[cell.row].height = calculated_height
                                            else:
                                                if new_sheet.row_dimensions[cell.row].height is None:
                                                     new_sheet.row_dimensions[cell.row].height = None
                                            
                            wb.remove(base_sheet)
                            wb_buffer = io.BytesIO()
                            wb.save(wb_buffer)
                            
                            safe_filename = str(trade_area).replace("/", "-").replace("\\", "-")
                            zip_file.writestr(f"{safe_filename}.xlsx", wb_buffer.getvalue())
                            progress_bar.progress((i + 1) / total_groups)
                    
                    status_text.empty()
                    progress_bar.empty()
                    
                    st.session_state.zip_data = zip_buffer.getvalue()
                    st.rerun()

        # STEP 4: The Export Section
        if st.session_state.zip_data is not None:
            st.success("Reports generated successfully!")
            st.download_button(
                label="Download All Reports (.zip)",
                data=st.session_state.zip_data,
                file_name="Trade_Area_Reports.zip",
                mime="application/zip",
                use_container_width=True
            )
            
            if st.button("Start Over"):
                st.session_state.zip_data = None
                st.rerun()
